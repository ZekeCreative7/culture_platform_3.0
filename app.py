import streamlit as st
import pandas as pd
import datetime
import io
import os
import altair as alt
import db
import parser as csv_parser

_DIR = os.path.dirname(os.path.abspath(__file__))
FAVICON_PATH = os.path.join(_DIR, "static", "assets", "favicon.png")
LOGO_PATH    = os.path.join(_DIR, "static", "assets", "lina_logo.png")

try:
    from PIL import Image as _PILImage
    _page_icon = _PILImage.open(FAVICON_PATH)
except Exception:
    _page_icon = "🏢"

st.set_page_config(
    page_title="Culture Platform 3.0",
    page_icon=_page_icon,
    layout="wide",
    initial_sidebar_state="expanded",
)

db.init_db()

# ── Constants ─────────────────────────────────────────────────────
PAGES = ["홈 대시보드", "세션 관리", "CSV 업로드", "변화량 조회", "경영진 보고"]
PHASES = ["사전", "중간", "사후"]

QUANT_LABELS = {
    "q1": "심리안전①", "q2": "심리안전②", "q3": "심리안전③",
    "q4": "사일로해소①", "q5": "사일로해소②", "q6": "사일로해소③",
    "q7": "회복/긴장",
    "q8": "전반 분위기",
}

SESSION_TYPE_INFO = {
    "팀빌딩":      {"target_weeks": 8,  "desc": "특정 팀 팀장 + 팀원 전체 참여"},
    "팀장":        {"target_weeks": 4,  "desc": "협업이 필요한 팀의 팀장들 참여"},
    "크로스펑셔널": {"target_weeks": 6,  "desc": "팀장 세션 팀장들이 차출한 팀원 2명씩 참여"},
}

CONTENT_TEMPLATES = {
    "팀빌딩": [
        {"content_name": "WOW세션",         "duration_min": 60},
        {"content_name": "명상세션",          "duration_min": 60},
        {"content_name": "커뮤니케이션세션",   "duration_min": 60},
        {"content_name": "간담회",            "duration_min": 60},
        {"content_name": "파트너요가",         "duration_min": 60},
        {"content_name": "에너지회복",         "duration_min": 60},
    ],
    "팀장": [
        {"content_name": "웰니스 + WOW세션", "duration_min": 120},
        {"content_name": "웰니스 + WOW세션", "duration_min": 120},
        {"content_name": "웰니스 + WOW세션", "duration_min": 120},
        {"content_name": "웰니스 + WOW세션", "duration_min": 120},
    ],
    "크로스펑셔널": [
        {"content_name": "크로스펑셔널 세션", "duration_min": 120},
        {"content_name": "크로스펑셔널 세션", "duration_min": 120},
        {"content_name": "크로스펑셔널 세션", "duration_min": 120},
        {"content_name": "크로스펑셔널 세션", "duration_min": 120},
        {"content_name": "크로스펑셔널 세션", "duration_min": 120},
        {"content_name": "크로스펑셔널 세션", "duration_min": 120},
    ],
}

TYPE_COLOR = {"팀빌딩": "#0066cc", "팀장": "#0a7f62", "크로스펑셔널": "#b66a00"}

# ── CSS ───────────────────────────────────────────────────────────
# Design tokens (from STYLE_GUIDE.md):
# Primary #5f6dee · Secondary #18b6aa · Accent #ffd166 · Risk #fb7185
# Strong #0b1020  · Muted #64748b   · Surface rgba(255,255,255,0.86)
# Line rgba(203,213,225,0.72)
st.markdown("""
<style>
:root {
    --cp-bg: #f5f5f7;
    --cp-surface: #ffffff;
    --cp-surface-soft: #fbfbfd;
    --cp-text: #1d1d1f;
    --cp-muted: #6e6e73;
    --cp-faint: #86868b;
    --cp-line: #d2d2d7;
    --cp-blue: #0066cc;
    --cp-green: #0a7f62;
    --cp-amber: #b66a00;
    --cp-red: #b42318;
}

html, body, [class*="css"] {
    font-family: -apple-system, BlinkMacSystemFont, "SF Pro Display",
                 "Apple SD Gothic Neo", "Malgun Gothic", sans-serif;
    -webkit-font-smoothing: antialiased;
}

.stApp { background: var(--cp-bg); }
.main .block-container {
    max-width: 1180px;
    padding-top: 2.1rem;
    padding-bottom: 4rem;
}

[data-testid="stSidebar"] {
    background: rgba(255,255,255,0.86) !important;
    border-right: 1px solid rgba(0,0,0,0.08) !important;
    min-width: 218px !important;
    max-width: 232px !important;
    backdrop-filter: saturate(180%) blur(20px);
}
[data-testid="stSidebar"] * { color: var(--cp-text) !important; }
[data-testid="stSidebar"] hr {
    border: none !important;
    border-top: 1px solid rgba(0,0,0,0.08) !important;
    margin: 12px 0 !important;
}
[data-testid="stSidebar"] .stRadio > label {
    font-size: 11px !important;
    letter-spacing: 0 !important;
    color: var(--cp-faint) !important;
    margin-bottom: 8px !important;
}
[data-testid="stSidebar"] .stRadio [data-testid="stMarkdownContainer"] p {
    font-size: 13px !important;
    font-weight: 500 !important;
    color: var(--cp-text) !important;
    line-height: 1.45;
}
[data-testid="stSidebar"] small {
    color: var(--cp-faint) !important;
    font-size: 12px !important;
}

h1 {
    color: var(--cp-text) !important;
    font-size: 2.25rem !important;
    font-weight: 700 !important;
    letter-spacing: 0 !important;
    margin: 0 0 0.25rem !important;
}
h2 {
    color: var(--cp-text) !important;
    font-size: 1.22rem !important;
    font-weight: 650 !important;
    letter-spacing: 0 !important;
    margin-top: 1.5rem !important;
}
h3 {
    color: var(--cp-text) !important;
    font-size: 1rem !important;
    font-weight: 650 !important;
    letter-spacing: 0 !important;
    text-transform: none !important;
    margin-top: 1.25rem !important;
}
p, li { color: var(--cp-text); font-size: 14px; }
.stCaption p { color: var(--cp-muted) !important; font-size: 13px !important; }

.cp-hero {
    background: radial-gradient(circle at 92% 15%, rgba(0,102,204,0.12), transparent 28%),
                linear-gradient(180deg, #ffffff 0%, #f5f5f7 100%);
    border: 1px solid rgba(0,0,0,0.06);
    border-radius: 8px;
    padding: 38px 40px 34px;
    margin: 0 0 22px;
    min-height: 224px;
    overflow: hidden;
}
.cp-eyebrow {
    color: var(--cp-muted);
    font-size: 13px;
    font-weight: 600;
    margin-bottom: 10px;
}
.cp-hero h1 {
    max-width: 720px;
    font-size: 3.05rem !important;
    line-height: 1.04 !important;
    margin-bottom: 12px !important;
}
.cp-hero p {
    max-width: 640px;
    color: var(--cp-muted);
    font-size: 1.16rem;
    line-height: 1.48;
    margin: 0;
}
.cp-hero-strip {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-top: 24px;
}
.cp-chip {
    display: inline-flex;
    align-items: center;
    min-height: 30px;
    padding: 5px 10px;
    border-radius: 999px;
    background: rgba(255,255,255,0.72);
    border: 1px solid rgba(0,0,0,0.08);
    color: var(--cp-muted);
    font-size: 12px;
    font-weight: 600;
}
.cp-section-title {
    display: flex;
    align-items: end;
    justify-content: space-between;
    gap: 16px;
    margin: 24px 0 10px;
}
.cp-section-title h2 {
    margin: 0 !important;
    font-size: 1.18rem !important;
}
.cp-section-title span {
    color: var(--cp-muted);
    font-size: 12px;
}

[data-testid="metric-container"] {
    background: var(--cp-surface);
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 8px;
    padding: 16px 18px !important;
    min-height: 104px;
    box-shadow: 0 1px 2px rgba(0,0,0,0.02);
}
[data-testid="stMetricLabel"] p {
    color: var(--cp-muted) !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    letter-spacing: 0 !important;
    text-transform: none !important;
}
[data-testid="stMetricValue"] {
    color: var(--cp-text) !important;
    font-size: 2.05rem !important;
    font-weight: 650 !important;
}
[data-testid="stMetricDelta"] { font-size: 12px !important; }

.stButton > button[kind="primary"] {
    background: var(--cp-blue) !important;
    color: white !important;
    border: 1px solid var(--cp-blue) !important;
    border-radius: 999px !important;
    font-size: 13px !important;
    font-weight: 600 !important;
    padding: 0.42rem 1.2rem !important;
    box-shadow: none !important;
    transition: background 0.12s ease, border-color 0.12s ease !important;
}
.stButton > button[kind="primary"]:hover {
    background: #0077ed !important;
    border-color: #0077ed !important;
}
.stButton > button:not([kind="primary"]) {
    background: var(--cp-surface) !important;
    color: var(--cp-blue) !important;
    border: 1px solid rgba(0,102,204,0.28) !important;
    border-radius: 999px !important;
    font-size: 13px !important;
    font-weight: 600 !important;
    padding: 0.36rem 1rem !important;
    box-shadow: none !important;
}
.stButton > button:not([kind="primary"]):hover {
    background: rgba(0,102,204,0.06) !important;
    border-color: rgba(0,102,204,0.45) !important;
}

[data-testid="stExpander"] {
    background: var(--cp-surface) !important;
    border: 1px solid rgba(0,0,0,0.08) !important;
    border-radius: 8px !important;
    box-shadow: none !important;
    overflow: hidden !important;
    margin-bottom: 10px !important;
}
[data-testid="stExpander"] summary {
    color: var(--cp-text) !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    padding: 13px 16px !important;
}
[data-testid="stExpander"] summary:hover { background: var(--cp-surface-soft) !important; }

.stTabs [data-baseweb="tab-list"] {
    background: transparent;
    border-bottom: 1px solid var(--cp-line);
    gap: 20px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 0 !important;
    font-size: 13px !important;
    font-weight: 600;
    color: var(--cp-muted);
    padding: 8px 0 10px !important;
    background: transparent;
    border: none;
}
.stTabs [aria-selected="true"] {
    color: var(--cp-text) !important;
    border-bottom: 2px solid var(--cp-text) !important;
    box-shadow: none !important;
}
.stTabs [data-baseweb="tab-panel"] { padding-top: 16px !important; }

[data-baseweb="input"], [data-baseweb="select"] > div {
    border-radius: 8px !important;
    border-color: var(--cp-line) !important;
    background: var(--cp-surface) !important;
    font-size: 13px !important;
}
[data-testid="stFileUploader"] section {
    border: 1px dashed var(--cp-line) !important;
    border-radius: 8px !important;
    background: var(--cp-surface) !important;
}
[data-testid="stFileUploader"] section:hover {
    border-color: var(--cp-blue) !important;
    background: #f7fbff !important;
}
[data-testid="stDataFrame"] > div {
    border-radius: 8px !important;
    overflow: hidden !important;
    border: 1px solid rgba(0,0,0,0.08) !important;
}
[data-testid="stAlert"] {
    border-radius: 8px !important;
    font-size: 13px !important;
}
hr {
    border: none !important;
    border-top: 1px solid var(--cp-line) !important;
    margin: 1.2rem 0 !important;
}

.badge {
    display: inline-flex;
    align-items: center;
    min-height: 24px;
    padding: 2px 8px;
    border-radius: 999px;
    font-size: 11px;
    font-weight: 650;
    letter-spacing: 0;
    line-height: 1.4;
    margin-right: 4px;
}
.badge-primary   { background: rgba(0,102,204,0.10); color: var(--cp-blue); }
.badge-secondary { background: rgba(10,127,98,0.10); color: var(--cp-green); }
.badge-accent    { background: rgba(182,106,0,0.12); color: var(--cp-amber); }
.badge-muted     { background: rgba(110,110,115,0.12); color: var(--cp-muted); }
.badge-risk      { background: rgba(180,35,24,0.10); color: var(--cp-red); }

.lina-card {
    background: var(--cp-surface);
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 8px;
    padding: 16px 18px;
    margin-bottom: 10px;
    box-shadow: 0 1px 2px rgba(0,0,0,0.02);
}
.lina-card-sm {
    background: var(--cp-surface);
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 8px;
    padding: 12px 14px;
    margin-bottom: 8px;
    font-size: 13px;
}
.cp-list-card {
    display: grid;
    gap: 6px;
    background: var(--cp-surface);
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 8px;
    padding: 15px 16px;
    margin-bottom: 10px;
}
.cp-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 12px;
}
.cp-kicker {
    color: var(--cp-muted);
    font-size: 12px;
    font-weight: 600;
}
.cp-title {
    color: var(--cp-text);
    font-size: 14px;
    font-weight: 650;
    line-height: 1.4;
}
.cp-sub {
    color: var(--cp-muted);
    font-size: 12px;
    line-height: 1.45;
}
.cp-type-grid {
    display: grid;
    grid-template-columns: 1fr;
    gap: 10px;
}
.cp-type-card {
    background: var(--cp-surface);
    border: 1px solid rgba(0,0,0,0.08);
    border-radius: 8px;
    padding: 15px 16px;
}
.cp-type-card .num {
    color: var(--cp-text);
    font-size: 1.75rem;
    font-weight: 650;
    line-height: 1.05;
    margin-top: 4px;
}
@media (max-width: 760px) {
    .main .block-container { padding-top: 1.2rem; }
    .cp-hero { padding: 28px 22px; min-height: auto; }
    .cp-hero h1 { font-size: 2.15rem !important; }
    .cp-hero p { font-size: 1rem; }
}
</style>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────
with st.sidebar:
    # Logo: small, not full-width
    if os.path.exists(LOGO_PATH):
        _, c, _ = st.columns([1, 6, 1])
        c.image(LOGO_PATH, width=110)
    st.markdown("---")
    page = st.radio("메뉴", PAGES, key="nav_page", label_visibility="visible")
    st.markdown("---")
    st.markdown(f"<small>{datetime.date.today().strftime('%Y.%m.%d')}</small>", unsafe_allow_html=True)


# ── Helpers ───────────────────────────────────────────────────────
def _get_session_status(schedule: list) -> tuple:
    today = str(datetime.date.today())
    confirmed = [r for r in schedule if r.get("scheduled_date")]
    if not confirmed:
        return "시작전", "badge-muted"
    past   = [r for r in confirmed if r["scheduled_date"] <= today]
    future = [r for r in confirmed if r["scheduled_date"] > today]
    unpln  = [r for r in schedule   if not r.get("scheduled_date")]
    if not past:
        return "시작전", "badge-accent"
    elif future or unpln:
        return "진행중", "badge-primary"
    else:
        return "완료", "badge-secondary"


def _phase_badges_html(phases_done: list) -> str:
    parts = []
    for ph in PHASES:
        if ph in phases_done:
            parts.append(f'<span class="badge badge-primary">{ph} 완료</span>')
        else:
            parts.append(f'<span class="badge badge-muted">{ph} 대기</span>')
    return " ".join(parts)


def _section_header(title: str, meta: str = ""):
    meta_html = f"<span>{meta}</span>" if meta else ""
    st.markdown(
        f"""
<div class="cp-section-title">
  <h2>{title}</h2>
  {meta_html}
</div>
""",
        unsafe_allow_html=True,
    )


def _session_label(s: dict) -> str:
    if s["type"] == "팀빌딩":
        return f"[{s['id']}] 팀빌딩 | {s['cohort']}기 | {s.get('team') or ''}"
    elif s["type"] == "팀장":
        teams = s.get("participating_teams") or []
        tstr  = ", ".join(teams) if teams else (s.get("bonbu") or "")
        return f"[{s['id']}] 팀장 | {s['cohort']}기 | {tstr}"
    else:
        return f"[{s['id']}] 크로스펑셔널 | {s['cohort']}기"


def _init_schedule_state(type_: str):
    tmpl   = CONTENT_TEMPLATES[type_]
    today  = datetime.date.today()
    st.session_state.ns_type  = type_
    st.session_state.ns_count = len(tmpl)
    for i, t in enumerate(tmpl):
        st.session_state[f"ns_confirmed_{i}"] = (i < 2)
        st.session_state[f"ns_date_{i}"]      = today + datetime.timedelta(weeks=i)
        st.session_state[f"ns_time_{i}"]      = "10:00"
        st.session_state[f"ns_content_{i}"]   = t["content_name"]
        st.session_state[f"ns_dur_{i}"]       = t["duration_min"]
        st.session_state[f"ns_note_{i}"]      = ""


def _make_excel_report(cohort: int, stats: list, sessions: list) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    stat_map = {s["phase"]: s for s in stats}
    pre  = stat_map.get("사전")
    post = stat_map.get("사후")
    pre_n  = pre["n"]  if pre  else 0
    post_n = post["n"] if post else 0
    masked = pre_n < 3 or post_n < 3

    rows = []
    for q, label in QUANT_LABELS.items():
        if masked:
            rows.append({"문항": label, "사전 평균": "N<3 (마스킹)", "사후 평균": "N<3 (마스킹)", "변화량(Δ)": "-"})
        else:
            pv = pre[f"{q}_avg"]  if pre  and pre.get(f"{q}_avg")  is not None else None
            qv = post[f"{q}_avg"] if post and post.get(f"{q}_avg") is not None else None
            delta = round(qv - pv, 2) if (pv is not None and qv is not None) else "-"
            rows.append({
                "문항": label,
                "사전 평균": round(pv, 2) if pv is not None else "-",
                "사후 평균": round(qv, 2) if qv is not None else "-",
                "변화량(Δ)": delta,
            })

    qual_rows = []
    for s in sessions:
        if s.get("cohort") != cohort:
            continue
        for ph in PHASES:
            for r in db.get_qualitative(s["id"], ph):
                if any(v for v in r.values() if v):
                    qual_rows.append({
                        "세션": _session_label(s),
                        "Phase": ph,
                        "좋았던 점": r.get("q10") or "",
                        "운영진에게": r.get("q11") or "",
                    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Sheet 1: summary
        df_s = pd.DataFrame(rows)
        df_s.to_excel(writer, sheet_name="집계 결과", index=False, startrow=3)
        ws = writer.sheets["집계 결과"]
        ws["A1"] = f"라이나생명 Culture Platform — {cohort}기 성과 보고"
        ws["A2"] = f"생성일: {datetime.date.today().strftime('%Y년 %m월 %d일')}  |  사전 N={pre_n}  사후 N={post_n}"
        ws["A1"].font = Font(name="Malgun Gothic", size=14, bold=True, color="5f6dee")
        ws["A2"].font = Font(name="Malgun Gothic", size=11, color="666666")
        hf  = Font(name="Malgun Gothic", bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="5f6dee", end_color="5f6dee", fill_type="solid")
        for col in range(1, 5):
            cell = ws.cell(row=4, column=col)
            cell.font  = hf
            cell.fill  = hfill
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row:
                cell.font = Font(name="Malgun Gothic", size=10)
                if cell.column == 4 and isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        cell.font = Font(name="Malgun Gothic", size=10, color="18b6aa", bold=True)
                    elif cell.value < 0:
                        cell.font = Font(name="Malgun Gothic", size=10, color="C0392B", bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 24

        # ── Sheet 2: qualitative
        if qual_rows:
            df_q = pd.DataFrame(qual_rows)
            df_q.to_excel(writer, sheet_name="정성 응답", index=False)
            wq = writer.sheets["정성 응답"]
            for col in wq.columns:
                wq.column_dimensions[col[0].column_letter].width = 40

    return output.getvalue()


# ══════════════════════════════════════════════════════════════════
# 홈 대시보드
# ══════════════════════════════════════════════════════════════════
if page == "홈 대시보드":
    today = datetime.date.today()
    week_end = today + datetime.timedelta(days=7)
    sessions  = db.list_sessions()
    all_sched = {s["id"]: db.get_schedule(s["id"]) for s in sessions}

    # Compute stats
    active_count   = sum(1 for s in sessions if _get_session_status(all_sched[s["id"]])[0] == "진행중")
    pending_alerts = []
    this_week_items = []

    for s in sessions:
        sched = all_sched[s["id"]]
        unconfirmed = [r for r in sched if not r.get("scheduled_date")]
        if unconfirmed:
            pending_alerts.append((s, len(unconfirmed)))
        for r in sched:
            if r.get("scheduled_date"):
                try:
                    d = datetime.date.fromisoformat(r["scheduled_date"])
                    if today <= d <= week_end:
                        this_week_items.append((d, r, s))
                except Exception:
                    pass

    this_week_items.sort(key=lambda x: x[0])

    cohort_count = len({s.get("cohort") for s in sessions if s.get("cohort") is not None})
    completion_ready = sum(
        1 for s in sessions
        if {"사전", "사후"}.issubset(set(db.get_phases_for_session(s["id"])))
    )

    st.markdown(f"""
<section class="cp-hero">
  <div class="cp-eyebrow">{today.strftime('%Y년 %m월 %d일')} · Culture Platform 3.0</div>
  <h1>변화적응 세션 운영을 한 화면에서 정리합니다.</h1>
  <p>이번 주 일정, 미정 회차, 설문 업로드 상태를 먼저 보여주고 보고 가능한 기수까지 빠르게 확인하는 운영자 콕핏입니다.</p>
  <div class="cp-hero-strip">
    <span class="cp-chip">{len(sessions)}개 세션</span>
    <span class="cp-chip">{cohort_count}개 기수</span>
    <span class="cp-chip">{completion_ready}개 보고 준비</span>
  </div>
</section>
""", unsafe_allow_html=True)

    # ── Metric row ─────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("전체 세션", len(sessions))
    c2.metric("진행중", active_count)
    c3.metric("이번 주 일정", len(this_week_items), delta=f"{len(this_week_items)}건")
    c4.metric("미정 알림", len(pending_alerts))

    col_left, col_right = st.columns([3, 2], gap="large")

    with col_left:
        # ── This week schedule ──────────────────────────────────
        _section_header("이번 주 일정", f"{len(this_week_items)}건")
        if not this_week_items:
            st.markdown('<div class="lina-card"><p style="color:#6e6e73;margin:0">이번 주 확정된 일정이 없습니다.</p></div>', unsafe_allow_html=True)
        else:
            for d, r, s in this_week_items:
                tc = TYPE_COLOR.get(s["type"], "#5f6dee")
                dow = ["월", "화", "수", "목", "금", "토", "일"][d.weekday()]
                st.markdown(f"""
<div class="cp-list-card" style="border-left:3px solid {tc};">
  <div class="cp-row">
    <div>
      <div class="cp-kicker">{d.strftime('%m/%d')} ({dow}) · {r.get('start_time','')} · {r.get('duration_min','')}분</div>
      <div class="cp-title">{r.get('content_name','')}</div>
    </div>
    <span class="badge badge-muted">{s['type']}</span>
  </div>
  <div class="cp-sub">{_session_label(s)}</div>
</div>""", unsafe_allow_html=True)

        # ── Upload status ───────────────────────────────────────
        _section_header("업로드 현황", "사전 · 중간 · 사후")
        if not sessions:
            st.markdown('<div class="lina-card"><p style="color:#6e6e73;margin:0">등록된 세션이 없습니다.</p></div>', unsafe_allow_html=True)
        else:
            for s in sessions:
                phases_done = db.get_phases_for_session(s["id"])
                tc = TYPE_COLOR.get(s["type"], "#5f6dee")
                badges = _phase_badges_html(phases_done)
                st.markdown(f"""
<div class="cp-list-card" style="border-left:3px solid {tc};">
  <div class="cp-title">{_session_label(s)}</div>
  <div>{badges}</div>
</div>""", unsafe_allow_html=True)

    with col_right:
        # ── Pending alerts ──────────────────────────────────────
        _section_header("미정 회차", f"{len(pending_alerts)}개 세션")
        if not pending_alerts:
            st.markdown('<div class="lina-card"><p style="color:#0a7f62;font-weight:650;margin:0">모든 회차 일정이 확정되었습니다.</p></div>', unsafe_allow_html=True)
        else:
            for s, cnt in pending_alerts:
                st.markdown(f"""
<div class="cp-list-card" style="border-left:3px solid #b66a00;">
  <div class="cp-kicker">미정 {cnt}회차</div>
  <div class="cp-title">{_session_label(s)}</div>
</div>""", unsafe_allow_html=True)

        # ── Session type summary ────────────────────────────────
        _section_header("세션 현황")
        type_cards = []
        for t, color in TYPE_COLOR.items():
            t_sessions = [s for s in sessions if s["type"] == t]
            in_prog = sum(1 for s in t_sessions if _get_session_status(all_sched[s["id"]])[0] == "진행중")
            type_cards.append(f"""
<div class="cp-type-card" style="border-left:3px solid {color};">
  <div class="cp-kicker">{t}</div>
  <div class="num">{len(t_sessions)}</div>
  <div class="cp-sub">진행중 {in_prog}개</div>
</div>""")
        st.markdown(f'<div class="cp-type-grid">{"".join(type_cards)}</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# 세션 관리
# ══════════════════════════════════════════════════════════════════
elif page == "세션 관리":
    st.markdown("# 세션 관리")

    # ── Pending unconfirmed alert ──────────────────────────────────
    sessions_all = db.list_sessions()
    alerts = []
    for s in sessions_all:
        sched = db.get_schedule(s["id"])
        unc = [r for r in sched if not r.get("scheduled_date")]
        if unc:
            alerts.append((s, unc))
    if alerts:
        total_unc = sum(len(u) for _, u in alerts)
        st.warning(f"미정 회차가 있는 세션 {len(alerts)}개 · 총 {total_unc}회차 날짜 미확정")

    with st.expander("새 세션 등록", expanded=True):

        # ── Session type ───────────────────────────────────────────
        type_ = st.selectbox(
            "세션 유형",
            list(SESSION_TYPE_INFO.keys()),
            format_func=lambda t: f"{t}  —  {SESSION_TYPE_INFO[t]['desc']}",
            key="new_type",
        )
        info = SESSION_TYPE_INFO[type_]
        st.caption(f"목표 기간: **{info['target_weeks']}주** (실제 일정은 스케줄로 결정)")

        if st.session_state.get("ns_type") != type_:
            _init_schedule_state(type_)

        # ── Cohort ─────────────────────────────────────────────────
        cohort = st.number_input("기수", min_value=1, step=1, value=1, key="new_cohort")
        st.divider()

        # ── Org info ───────────────────────────────────────────────
        divisions   = db.get_org_units("division")
        div_options = {u["name"]: u["id"] for u in divisions}

        if type_ == "팀빌딩":
            st.markdown("**조직 정보** — 부문 > 본부 > 팀")
            c1, c2, c3 = st.columns(3)
            div_name   = c1.selectbox("부문", ["(선택)"] + list(div_options.keys()), key="tb_div")
            div_id     = div_options.get(div_name)
            hq_opts    = {u["name"]: u["id"] for u in db.get_org_units("hq", div_id)} if div_id else {}
            bonbu_name = c2.selectbox("본부", ["(선택)"] + list(hq_opts.keys()), key="tb_hq")
            hq_id      = hq_opts.get(bonbu_name)
            team_opts  = {u["name"]: u["id"] for u in db.get_org_units("team", hq_id)} if hq_id else {}
            team_name  = c3.selectbox("팀", ["(선택)"] + list(team_opts.keys()), key="tb_team")
            division   = div_name   if div_name   != "(선택)" else None
            bonbu      = bonbu_name if bonbu_name != "(선택)" else None
            team       = team_name  if team_name  != "(선택)" else None
            participating_teams = None
            linked_session_id   = None

        elif type_ == "팀장":
            st.markdown("**조직 정보** — 협업이 필요한 팀들")
            c1, c2 = st.columns(2)
            div_name   = c1.selectbox("부문", ["(선택)"] + list(div_options.keys()), key="ls_div")
            div_id     = div_options.get(div_name)
            hq_opts    = {u["name"]: u["id"] for u in db.get_org_units("hq", div_id)} if div_id else {}
            bonbu_name = c2.selectbox("본부", ["(선택)"] + list(hq_opts.keys()), key="ls_hq")
            hq_id      = hq_opts.get(bonbu_name)
            all_teams  = [u["name"] for u in db.get_org_units("team", hq_id)] if hq_id else []
            sel_teams  = st.multiselect("참여 팀 선택", all_teams, key="ls_teams")
            division   = div_name   if div_name   != "(선택)" else None
            bonbu      = bonbu_name if bonbu_name != "(선택)" else None
            team       = None
            participating_teams = sel_teams
            linked_session_id   = None

        else:  # 크로스펑셔널
            st.markdown("**연결된 팀장 세션**")
            leader_sessions = db.list_sessions_by_type("팀장")
            if not leader_sessions:
                st.warning("먼저 팀장 세션을 등록하세요.")
                linked_session_id = None
            else:
                ls_opts           = {_session_label(s): s["id"] for s in leader_sessions}
                linked_label      = st.selectbox("팀장 세션 선택", list(ls_opts.keys()), key="cf_session")
                linked_session_id = ls_opts[linked_label]
            division = bonbu = team = None
            participating_teams = None

        st.divider()

        # ── Schedule builder ───────────────────────────────────────
        count = st.session_state.get("ns_count", len(CONTENT_TEMPLATES[type_]))
        st.markdown(f"**세션 스케줄** — {count}회차")
        st.caption("처음 2회차 날짜를 확정하고, 나머지는 진행하면서 추가하세요.")

        hdr = st.columns([1, 1, 2, 1, 3, 2, 2])
        for h, t in zip(hdr, ["회차", "날짜 확정", "날짜", "시작", "콘텐츠명", "소요(분)", "메모"]):
            h.markdown(f"**{t}**")

        for i in range(count):
            cols = st.columns([1, 1, 2, 1, 3, 2, 2])
            cols[0].markdown(f"**{i+1}회**")
            confirmed = cols[1].checkbox(f"{i+1}회차 날짜 확정", key=f"ns_confirmed_{i}", label_visibility="collapsed")
            if confirmed:
                cols[2].date_input(f"{i+1}회차 날짜", key=f"ns_date_{i}", label_visibility="collapsed")
            else:
                cols[2].markdown("<span style='color:#B0BAC9;font-size:13px'>미정</span>", unsafe_allow_html=True)
            cols[3].text_input(f"{i+1}회차 시작 시간", key=f"ns_time_{i}", label_visibility="collapsed")
            cols[4].text_input(f"{i+1}회차 콘텐츠명", key=f"ns_content_{i}", label_visibility="collapsed")
            cols[5].number_input(f"{i+1}회차 소요 시간", key=f"ns_dur_{i}", min_value=30, step=30, label_visibility="collapsed")
            cols[6].text_input(f"{i+1}회차 메모", key=f"ns_note_{i}", label_visibility="collapsed")

        col_add, col_del, _ = st.columns([2, 2, 6])
        if col_add.button("회차 추가"):
            last = count - 1
            ni   = count
            st.session_state[f"ns_confirmed_{ni}"] = False
            st.session_state[f"ns_date_{ni}"]      = datetime.date.today()
            st.session_state[f"ns_time_{ni}"]      = st.session_state.get(f"ns_time_{last}", "10:00")
            st.session_state[f"ns_content_{ni}"]   = st.session_state.get(f"ns_content_{last}", "")
            st.session_state[f"ns_dur_{ni}"]       = st.session_state.get(f"ns_dur_{last}", 60)
            st.session_state[f"ns_note_{ni}"]      = ""
            st.session_state.ns_count = count + 1
            st.rerun()

        if count > 1 and col_del.button("마지막 회차 삭제"):
            last = count - 1
            for k in ["confirmed", "date", "time", "content", "dur", "note"]:
                st.session_state.pop(f"ns_{k}_{last}", None)
            st.session_state.ns_count = count - 1
            st.rerun()

        st.divider()

        if st.button("세션 등록", type="primary"):
            err = None
            if type_ == "팀빌딩" and not team:
                err = "팀을 선택하세요."
            elif type_ == "팀장" and not sel_teams:
                err = "참여 팀을 선택하세요."
            elif type_ == "크로스펑셔널" and not linked_session_id:
                err = "연결할 팀장 세션을 선택하세요."

            if err:
                st.error(err)
            else:
                sid = db.add_session(
                    type_=type_, cohort=int(cohort),
                    division=division, bonbu=bonbu, team=team,
                    participating_teams=participating_teams,
                    linked_session_id=linked_session_id,
                )
                items = []
                for i in range(st.session_state.ns_count):
                    cfm = st.session_state.get(f"ns_confirmed_{i}", False)
                    items.append({
                        "seq":            i + 1,
                        "scheduled_date": str(st.session_state[f"ns_date_{i}"]) if cfm else None,
                        "start_time":     st.session_state.get(f"ns_time_{i}", "10:00"),
                        "duration_min":   st.session_state.get(f"ns_dur_{i}", 60),
                        "content_name":   st.session_state.get(f"ns_content_{i}", ""),
                        "content_note":   st.session_state.get(f"ns_note_{i}", ""),
                    })
                db.add_schedule_items(sid, items)
                st.success(f"세션 등록 완료 (ID: {sid})")
                st.session_state["_last_sid"] = sid
                _init_schedule_state(type_)

                # Post-registration navigation
                c_a, c_b, _ = st.columns([2, 2, 4])
                if c_a.button("CSV 업로드로 이동", type="primary", key="goto_csv_post"):
                    st.session_state["nav_page"] = "CSV 업로드"
                    st.rerun()
                if c_b.button("새 세션 추가", key="new_sess_post"):
                    st.rerun()

    # ── Session list (tabbed by type) ──────────────────────────────
    st.markdown("### 등록된 세션 목록")
    sessions = db.list_sessions()
    if not sessions:
        st.info("등록된 세션이 없습니다.")
    else:
        tab_labels = ["전체"] + list(SESSION_TYPE_INFO.keys())
        tabs = st.tabs(tab_labels)
        for ti, tab in enumerate(tabs):
            with tab:
                filter_type = None if ti == 0 else tab_labels[ti]
                filtered    = sessions if filter_type is None else [s for s in sessions if s["type"] == filter_type]
                if not filtered:
                    st.info("해당 유형의 세션이 없습니다.")
                    continue

                for s in filtered:
                    schedule       = db.get_schedule(s["id"])
                    phases_done    = db.get_phases_for_session(s["id"])
                    status, scls   = _get_session_status(schedule)
                    confirmed_cnt  = len([r for r in schedule if r.get("scheduled_date")])
                    pending_cnt    = len([r for r in schedule if not r.get("scheduled_date")])
                    tc             = TYPE_COLOR.get(s["type"], "#5f6dee")
                    badges_html    = _phase_badges_html(phases_done)

                    with st.expander(
                        f"{_session_label(s)}  ·  확정 {confirmed_cnt}회 / 미정 {pending_cnt}회",
                        expanded=False
                    ):
                        # Status + phase badges
                        st.markdown(
                            f'<span class="badge {scls}">{status}</span>&nbsp;&nbsp;{badges_html}',
                            unsafe_allow_html=True,
                        )
                        st.markdown("<br>", unsafe_allow_html=True)

                        # Schedule table
                        if schedule:
                            tbl = [{
                                "회차":     f"{r['seq']}회",
                                "날짜":     r["scheduled_date"] or "미정",
                                "시작":     r["start_time"] or "",
                                "콘텐츠":   r["content_name"] or "",
                                "소요(분)": r["duration_min"] or "",
                                "메모":     r["content_note"] or "",
                                "상태":     r["status"],
                            } for r in schedule]
                            st.dataframe(pd.DataFrame(tbl), use_container_width=True, hide_index=True)

                        # Date confirmation for unconfirmed items
                        pending_items = [r for r in schedule if not r.get("scheduled_date")]
                        if pending_items:
                            st.markdown("**날짜 확정 추가**")
                            for r in pending_items:
                                cc1, cc2, cc3 = st.columns([1, 2, 2])
                                cc1.write(f"{r['seq']}회차")
                                new_date = cc2.date_input("날짜", key=f"upd_date_{r['id']}", label_visibility="collapsed")
                                if cc3.button("확정", key=f"upd_btn_{r['id']}"):
                                    db.update_schedule_item(r["id"], scheduled_date=str(new_date), status="confirmed")
                                    st.rerun()


# ══════════════════════════════════════════════════════════════════
# CSV 업로드
# ══════════════════════════════════════════════════════════════════
elif page == "CSV 업로드":
    st.markdown("# CSV 업로드")

    sessions = db.list_sessions()
    if not sessions:
        st.warning("먼저 세션을 등록하세요.")
        st.stop()

    session_options  = {_session_label(s): s["id"] for s in sessions}
    selected_label   = st.selectbox("세션 선택", list(session_options.keys()))
    session_id       = session_options[selected_label]
    selected_session = next(s for s in sessions if s["id"] == session_id)

    # Phase upload status
    phases_done = db.get_phases_for_session(session_id)
    st.markdown(
        f"**업로드 현황** &nbsp;&nbsp; {_phase_badges_html(phases_done)}",
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    phase    = st.selectbox("Phase", PHASES)
    uploaded = st.file_uploader("구글 폼 CSV 업로드", type=["csv"])

    if uploaded:
        csv_bytes = uploaded.read()
        rows, errors = csv_parser.validate_and_parse(csv_bytes, session_id, phase)
        if errors:
            for e in errors:
                st.error(e)
        else:
            st.success(f"검증 통과 · {len(rows)}행")
            st.dataframe(pd.DataFrame(rows), use_container_width=True)
            if st.button("적재 확정", type="primary"):
                db.save_responses(rows)
                st.success(f"{len(rows)}행 저장 완료")
                st.balloons()
                st.rerun()


# ══════════════════════════════════════════════════════════════════
# 변화량 조회
# ══════════════════════════════════════════════════════════════════
elif page == "변화량 조회":
    st.markdown("# 변화량 조회")

    all_cohorts = db.get_all_cohorts()
    if not all_cohorts:
        st.info("적재된 응답 데이터가 없습니다.")
        st.stop()

    cohort   = st.selectbox("기수 선택", all_cohorts)
    stats    = db.get_cohort_stats(cohort)
    if not stats:
        st.info("해당 기수의 집계 데이터가 없습니다.")
        st.stop()

    stat_map = {s["phase"]: s for s in stats}
    pre      = stat_map.get("사전")
    mid      = stat_map.get("중간")
    post     = stat_map.get("사후")

    # Metrics
    cc = st.columns(3)
    cc[0].metric("사전 N", pre["n"]  if pre  else "—")
    cc[1].metric("중간 N", mid["n"]  if mid  else "—")
    cc[2].metric("사후 N", post["n"] if post else "—")

    st.markdown("---")

    # ── Chart ──────────────────────────────────────────────────────
    chart_rows = []
    for q, label in QUANT_LABELS.items():
        for ph, s in [("사전", pre), ("중간", mid), ("사후", post)]:
            if s and s.get(f"{q}_avg") is not None:
                chart_rows.append({"문항": label, "구분": ph, "평균": round(s[f"{q}_avg"], 2)})

    if chart_rows:
        df_chart = pd.DataFrame(chart_rows)
        color_scale = alt.Scale(
            domain=["사전", "중간", "사후"],
            range=["#a7c7e7", "#d6a24a", "#0066cc"]
        )
        chart = (
            alt.Chart(df_chart)
            .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
            .encode(
                x=alt.X("문항:N", sort=None, axis=alt.Axis(labelAngle=-30, labelFontSize=11)),
                y=alt.Y("평균:Q", scale=alt.Scale(domain=[0, 5]),
                        axis=alt.Axis(tickCount=6)),
                color=alt.Color("구분:N", scale=color_scale,
                                legend=alt.Legend(orient="top", title=None)),
                xOffset="구분:N",
                tooltip=["문항", "구분", "평균"],
            )
            .properties(height=320, title=alt.TitleParams("사전 · 중간 · 사후 변화량", fontSize=14, fontWeight=600))
        )
        st.altair_chart(chart, use_container_width=True)

    st.markdown("---")

    # ── Table ──────────────────────────────────────────────────────
    table_rows = []
    for q, label in QUANT_LABELS.items():
        pv = round(pre[f"{q}_avg"],  2) if pre  and pre.get(f"{q}_avg")  is not None else None
        mv = round(mid[f"{q}_avg"],  2) if mid  and mid.get(f"{q}_avg")  is not None else None
        qv = round(post[f"{q}_avg"], 2) if post and post.get(f"{q}_avg") is not None else None
        delta = round(qv - pv, 2) if (pv is not None and qv is not None) else None
        table_rows.append({
            "문항":      label,
            "사전 평균": pv  if pv  is not None else "—",
            "중간 평균": mv  if mv  is not None else "—",
            "사후 평균": qv  if qv  is not None else "—",
            "변화량(Δ)": delta if delta is not None else "—",
        })
    st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

    # ── Qualitative responses ───────────────────────────────────────
    sessions   = db.list_sessions()
    session_ids = [s["id"] for s in sessions if s.get("cohort") == cohort]
    for sid in session_ids:
        for phase_label in PHASES:
            qual  = db.get_qualitative(sid, phase_label)
            texts = [r for r in qual if any(v for v in r.values() if v)]
            if not texts:
                continue
            st.markdown(f"#### 정성 응답 — {phase_label}")
            for r in texts:
                if r.get("q9"):  st.markdown(f"- **[기대]** {r['q9']}")
                if r.get("q10"): st.markdown(f"- **[좋았던 점]** {r['q10']}")
                if r.get("q11"): st.markdown(f"- **[운영진에게]** {r['q11']}")


# ══════════════════════════════════════════════════════════════════
# 경영진 보고
# ══════════════════════════════════════════════════════════════════
elif page == "경영진 보고":
    st.markdown("# 경영진 보고")
    st.caption("응답 인원 N < 3인 셀은 개인 식별 방지를 위해 마스킹됩니다.")

    all_cohorts = db.get_all_cohorts()
    if not all_cohorts:
        st.info("적재된 응답 데이터가 없습니다.")
        st.stop()

    cohort   = st.selectbox("기수 선택", all_cohorts)
    stats    = db.get_cohort_stats(cohort)
    stat_map = {s["phase"]: s for s in stats}
    pre      = stat_map.get("사전")
    post     = stat_map.get("사후")
    pre_n    = pre["n"]  if pre  else 0
    post_n   = post["n"] if post else 0

    c1, c2 = st.columns(2)
    c1.metric("사전 응답 N", pre_n)
    c2.metric("사후 응답 N", post_n)

    MASK = "N<3 (마스킹)"
    table_rows = []
    for q, label in QUANT_LABELS.items():
        if pre_n < 3 or post_n < 3:
            table_rows.append({"문항": label, "사전 평균": MASK, "사후 평균": MASK, "변화량(Δ)": MASK})
            continue
        pv = pre[f"{q}_avg"]  if pre  and pre.get(f"{q}_avg")  is not None else None
        qv = post[f"{q}_avg"] if post and post.get(f"{q}_avg") is not None else None
        delta = round(qv - pv, 2) if (pv is not None and qv is not None) else "-"
        table_rows.append({
            "문항":      label,
            "사전 평균": round(pv, 2) if pv is not None else "-",
            "사후 평균": round(qv, 2) if qv is not None else "-",
            "변화량(Δ)": delta,
        })

    st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── Chart ──────────────────────────────────────────────────────
    if pre_n >= 3 and post_n >= 3:
        chart_rows = []
        for q, label in QUANT_LABELS.items():
            for ph, s in [("사전", pre), ("사후", post)]:
                if s and s.get(f"{q}_avg") is not None:
                    chart_rows.append({"문항": label, "구분": ph, "평균": round(s[f"{q}_avg"], 2)})
        if chart_rows:
            df_chart = pd.DataFrame(chart_rows)
            chart = (
                alt.Chart(df_chart)
                .mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
                .encode(
                    x=alt.X("문항:N", sort=None, axis=alt.Axis(labelAngle=-30)),
                    y=alt.Y("평균:Q", scale=alt.Scale(domain=[0, 5])),
                    color=alt.Color("구분:N",
                                    scale=alt.Scale(domain=["사전", "사후"],
                                                    range=["#a7c7e7", "#0066cc"]),
                                    legend=alt.Legend(orient="top", title=None)),
                    xOffset="구분:N",
                    tooltip=["문항", "구분", "평균"],
                )
                .properties(height=300)
            )
            st.altair_chart(chart, use_container_width=True)

    st.markdown("---")

    # ── Export ─────────────────────────────────────────────────────
    st.markdown("### 보고서 내보내기")
    sessions_all = db.list_sessions()
    xlsx_data = _make_excel_report(cohort, stats, sessions_all)
    st.download_button(
        label="Excel 보고서 다운로드",
        data=xlsx_data,
        file_name=f"culture_report_{cohort}기_{datetime.date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
