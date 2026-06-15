"""
조직도 엑셀 → DB 적재 스크립트
실행: python3 load_org.py
"""
import os
import openpyxl
import db

EXCEL_PATH = os.path.expanduser(
    "~/Downloads/lina_organization_template (1).xlsx"
)


def load():
    db.init_db()

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Organization"]
    headers = [c.value for c in ws[1]]

    units = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] != "unit":
            continue
        d = dict(zip(headers, row))
        units.append((
            d["id"],
            d["level"],
            d.get("parentId") or None,
            d["name"],
            d.get("leader") or None,
            d.get("leaderTitle") or None,
            d.get("leaderRole") or None,
        ))

    conn = db.get_conn()
    conn.execute("DELETE FROM org_unit")
    conn.executemany(
        """INSERT INTO org_unit (id, level, parent_id, name, leader, leader_title, leader_role)
           VALUES (?,?,?,?,?,?,?)""",
        units,
    )
    conn.commit()
    conn.close()

    print(f"조직도 적재 완료: {len(units)}개 unit")
    by_level = {}
    for u in units:
        by_level[u[1]] = by_level.get(u[1], 0) + 1
    for level, cnt in sorted(by_level.items()):
        print(f"  {level}: {cnt}개")


if __name__ == "__main__":
    load()
